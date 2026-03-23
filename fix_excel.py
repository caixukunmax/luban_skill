#!/usr/bin/env python3
import openpyxl

# 修复道具表
wb = openpyxl.load_workbook(r'c:\code\luban_skill\tables\datas\#Item-道具表.xlsx')
sheet = wb.active

# 检查所有数据行，第一列应该是空的
fixed_rows = []
for row_num in range(6, sheet.max_row + 1):  # 从第6行开始（数据行）
    cell = sheet.cell(row=row_num, column=1)
    if cell.value and str(cell.value).startswith('##'):
        # 这是数据行被错误标记
        fixed_rows.append(row_num)
        cell.value = None

wb.save(r'c:\code\luban_skill\tables\datas\#Item-道具表.xlsx')
wb.close()

if fixed_rows:
    print(f'Fixed rows: {fixed_rows}')
else:
    print('No rows need fixing')
