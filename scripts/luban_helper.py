#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Luban 配置编辑器辅助脚本
用于 AI 操作 Luban 配置表、枚举、Bean 等
"""

import argparse
import json
import hashlib
import os
import sys
from pathlib import Path
from typing import Optional, List, Dict, Any

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    print("错误: 请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)


class LubanConfigHelper:
    """Luban 配置辅助类"""
    
    def __init__(self, data_dir: str, cache_dir: str = ".luban_cache"):
        self.data_dir = Path(data_dir)
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(exist_ok=True)
        
        # 定义文件路径
        self.enums_file = self.data_dir / "__enums__.xlsx"
        self.beans_file = self.data_dir / "__beans__.xlsx"
        self.tables_file = self.data_dir / "__tables__.xlsx"
    
    # ==================== 枚举操作 ====================
    
    def list_enums(self) -> List[Dict[str, Any]]:
        """列出所有枚举"""
        if not self.enums_file.exists():
            return []
        
        wb = openpyxl.load_workbook(self.enums_file)
        sheet = wb.active
        
        enums = []
        current_enum = None
        
        for row in sheet.iter_rows(min_row=4, values_only=True):
            full_name = row[1]  # B列
            
            if full_name:  # 新枚举定义
                if current_enum:
                    enums.append(current_enum)
                current_enum = {
                    "full_name": full_name,
                    "flags": row[2],
                    "unique": row[3],
                    "comment": row[6] if len(row) > 6 else "",
                    "items": []
                }
                # 检查同一行是否有第一个枚举项
                first_item_name = row[7] if len(row) > 7 else None
                if first_item_name:
                    current_enum["items"].append({
                        "name": first_item_name,
                        "alias": row[8] if len(row) > 8 else "",
                        "value": row[9] if len(row) > 9 else None,
                        "comment": row[10] if len(row) > 10 else ""
                    })
            elif current_enum:  # 枚举项
                # *items 列开始 (H列开始，索引7)
                item_name = row[7] if len(row) > 7 else None
                if item_name:
                    current_enum["items"].append({
                        "name": item_name,
                        "alias": row[8] if len(row) > 8 else "",
                        "value": row[9] if len(row) > 9 else None,
                        "comment": row[10] if len(row) > 10 else ""
                    })
        
        if current_enum:
            enums.append(current_enum)
        
        wb.close()
        return enums
    
    def get_enum(self, enum_name: str) -> Optional[Dict[str, Any]]:
        """获取指定枚举"""
        enums = self.list_enums()
        for enum in enums:
            if enum["full_name"] == enum_name or enum["full_name"].endswith("." + enum_name):
                return enum
        return None
    
    def add_enum(self, full_name: str, items: List[Dict], flags: bool = False, 
                 unique: bool = True, comment: str = "") -> bool:
        """新增枚举"""
        if not self.enums_file.exists():
            print(f"错误: 文件不存在 {self.enums_file}")
            return False
        
        # 检查是否已存在
        existing = self.get_enum(full_name)
        if existing:
            print(f"错误: 枚举 {full_name} 已存在")
            return False
        
        wb = openpyxl.load_workbook(self.enums_file)
        sheet = wb.active
        
        # 找到最后一行
        last_row = sheet.max_row
        
        # 添加枚举定义行
        row_num = last_row + 1
        sheet.cell(row=row_num, column=2, value=full_name)  # full_name
        sheet.cell(row=row_num, column=3, value=flags)       # flags
        sheet.cell(row=row_num, column=4, value=unique)      # unique
        sheet.cell(row=row_num, column=7, value=comment)     # comment
        
        # 添加第一个枚举项
        if items:
            item = items[0]
            sheet.cell(row=row_num, column=8, value=item.get("name"))
            sheet.cell(row=row_num, column=9, value=item.get("alias", ""))
            sheet.cell(row=row_num, column=10, value=item.get("value"))
            sheet.cell(row=row_num, column=11, value=item.get("comment", ""))
        
        # 添加剩余枚举项
        for i, item in enumerate(items[1:], start=1):
            row_num = last_row + 1 + i
            sheet.cell(row=row_num, column=8, value=item.get("name"))
            sheet.cell(row=row_num, column=9, value=item.get("alias", ""))
            sheet.cell(row=row_num, column=10, value=item.get("value"))
            sheet.cell(row=row_num, column=11, value=item.get("comment", ""))
        
        wb.save(self.enums_file)
        wb.close()
        
        print(f"✓ 已添加枚举: {full_name}")
        return True
    
    def delete_enum(self, enum_name: str) -> bool:
        """删除枚举"""
        if not self.enums_file.exists():
            print(f"错误: 文件不存在 {self.enums_file}")
            return False
        
        wb = openpyxl.load_workbook(self.enums_file)
        sheet = wb.active
        
        # 找到枚举的起始和结束行
        start_row = None
        end_row = None
        
        for i, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            full_name = row[1]
            
            if full_name == enum_name or (full_name and full_name.endswith("." + enum_name)):
                start_row = i
            elif start_row and full_name:  # 遇到下一个枚举
                end_row = i - 1
                break
        
        if not start_row:
            print(f"错误: 未找到枚举 {enum_name}")
            wb.close()
            return False
        
        if not end_row:
            end_row = sheet.max_row
        
        # 从后往前删除行
        for row_num in range(end_row, start_row - 1, -1):
            sheet.delete_rows(row_num)
        
        wb.save(self.enums_file)
        wb.close()
        
        print(f"✓ 已删除枚举: {enum_name}")
        return True
    
    # ==================== Bean 操作 ====================
    
    def list_beans(self) -> List[Dict[str, Any]]:
        """列出所有 Bean"""
        if not self.beans_file.exists():
            return []
        
        wb = openpyxl.load_workbook(self.beans_file)
        sheet = wb.active
        
        beans = []
        current_bean = None
        
        for row in sheet.iter_rows(min_row=4, values_only=True):
            full_name = row[1]  # B列
            
            if full_name:  # 新 Bean 定义
                if current_bean:
                    beans.append(current_bean)
                current_bean = {
                    "full_name": full_name,
                    "parent": row[2] if len(row) > 2 else "",
                    "value_type": row[3] if len(row) > 3 else "",
                    "alias": row[4] if len(row) > 4 else "",
                    "sep": row[5] if len(row) > 5 else "",
                    "comment": row[6] if len(row) > 6 else "",
                    "group": row[8] if len(row) > 8 else "",
                    "fields": []
                }
                # 检查同一行是否有第一个字段
                first_field_name = row[9] if len(row) > 9 else None
                if first_field_name:
                    current_bean["fields"].append({
                        "name": first_field_name,
                        "alias": row[10] if len(row) > 10 else "",
                        "type": row[11] if len(row) > 11 else "",
                        "group": row[12] if len(row) > 12 else "",
                        "comment": row[13] if len(row) > 13 else ""
                    })
            elif current_bean:  # 字段行
                # *fields 列开始 (I列开始，索引9)
                field_name = row[9] if len(row) > 9 else None
                if field_name:
                    current_bean["fields"].append({
                        "name": field_name,
                        "alias": row[10] if len(row) > 10 else "",
                        "type": row[11] if len(row) > 11 else "",
                        "group": row[12] if len(row) > 12 else "",
                        "comment": row[13] if len(row) > 13 else ""
                    })
        
        if current_bean:
            beans.append(current_bean)
        
        wb.close()
        return beans
    
    def get_bean(self, bean_name: str) -> Optional[Dict[str, Any]]:
        """获取指定 Bean"""
        beans = self.list_beans()
        for bean in beans:
            if bean["full_name"] == bean_name or bean["full_name"].endswith("." + bean_name):
                return bean
        return None
    
    def add_bean(self, full_name: str, fields: List[Dict], parent: str = "",
                 comment: str = "", alias: str = "") -> bool:
        """新增 Bean"""
        if not self.beans_file.exists():
            print(f"错误: 文件不存在 {self.beans_file}")
            return False
        
        # 检查是否已存在
        existing = self.get_bean(full_name)
        if existing:
            print(f"错误: Bean {full_name} 已存在")
            return False
        
        wb = openpyxl.load_workbook(self.beans_file)
        sheet = wb.active
        
        # 找到最后一行
        last_row = sheet.max_row
        
        # 添加 Bean 定义行
        row_num = last_row + 1
        sheet.cell(row=row_num, column=2, value=full_name)  # full_name
        sheet.cell(row=row_num, column=3, value=parent)     # parent
        sheet.cell(row=row_num, column=5, value=alias)      # alias
        sheet.cell(row=row_num, column=7, value=comment)    # comment
        
        # 添加第一个字段
        if fields:
            field = fields[0]
            sheet.cell(row=row_num, column=10, value=field.get("name"))
            sheet.cell(row=row_num, column=11, value=field.get("alias", ""))
            sheet.cell(row=row_num, column=12, value=field.get("type"))
            sheet.cell(row=row_num, column=14, value=field.get("comment", ""))
        
        # 添加剩余字段
        for i, field in enumerate(fields[1:], start=1):
            row_num = last_row + 1 + i
            sheet.cell(row=row_num, column=10, value=field.get("name"))
            sheet.cell(row=row_num, column=11, value=field.get("alias", ""))
            sheet.cell(row=row_num, column=12, value=field.get("type"))
            sheet.cell(row=row_num, column=14, value=field.get("comment", ""))
        
        wb.save(self.beans_file)
        wb.close()
        
        print(f"✓ 已添加 Bean: {full_name}")
        return True
    
    def delete_bean(self, bean_name: str) -> bool:
        """删除 Bean"""
        if not self.beans_file.exists():
            print(f"错误: 文件不存在 {self.beans_file}")
            return False
        
        wb = openpyxl.load_workbook(self.beans_file)
        sheet = wb.active
        
        # 找到 Bean 的起始和结束行
        start_row = None
        end_row = None
        
        for i, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            full_name = row[1]
            
            if full_name == bean_name or (full_name and full_name.endswith("." + bean_name)):
                start_row = i
            elif start_row and full_name:  # 遇到下一个 Bean
                end_row = i - 1
                break
        
        if not start_row:
            print(f"错误: 未找到 Bean {bean_name}")
            wb.close()
            return False
        
        if not end_row:
            end_row = sheet.max_row
        
        # 从后往前删除行
        for row_num in range(end_row, start_row - 1, -1):
            sheet.delete_rows(row_num)
        
        wb.save(self.beans_file)
        wb.close()
        
        print(f"✓ 已删除 Bean: {bean_name}")
        return True
    
    # ==================== 表操作 ====================
    
    def list_tables(self) -> List[Dict[str, Any]]:
        """列出所有表"""
        if not self.tables_file.exists():
            return []
        
        wb = openpyxl.load_workbook(self.tables_file)
        sheet = wb.active
        
        tables = []
        
        for row in sheet.iter_rows(min_row=4, values_only=True):
            full_name = row[1]  # B列
            if full_name:
                tables.append({
                    "full_name": full_name,
                    "value": row[2] if len(row) > 2 else "",
                    "input": row[3] if len(row) > 3 else "",
                    "mode": row[4] if len(row) > 4 else "",
                    "comment": row[5] if len(row) > 5 else ""
                })
        
        wb.close()
        return tables
    
    def get_table_data(self, table_name: str) -> Optional[Dict[str, Any]]:
        """获取表的数据（从 Excel 文件）"""
        # 查找匹配的表定义
        tables = self.list_tables()
        target_table = None
        for t in tables:
            if t["full_name"] == table_name or t["full_name"].endswith("." + table_name):
                target_table = t
                break
        
        if not target_table:
            return None
        
        # 解析 input 字段找到对应的 Excel 文件
        input_val = target_table.get("input", "")
        if not input_val or isinstance(input_val, bool):
            # 如果是布尔值 true，尝试直接从 mode 字段找文件名
            mode_str = target_table.get("mode", "")
            if mode_str and isinstance(mode_str, str):
                excel_files = list(self.data_dir.glob(f"**/{mode_str}"))
                if excel_files:
                    try:
                        return self._parse_excel_data(excel_files[0])
                    except Exception as e:
                        print(f"读取 {excel_files[0]} 失败：{e}")
            return None
        
        # 简单的文件名提取（去掉路径和@符号后的内容）
        file_part = str(input_val).split("@")[0] if "@" in str(input_val) else str(input_val)
        excel_files = list(self.data_dir.glob("**/*.xlsx"))
        
        for excel_file in excel_files:
            if file_part in str(excel_file) or file_part == excel_file.name:
                try:
                    return self._parse_excel_data(excel_file)
                except Exception as e:
                    print(f"读取 {excel_file} 失败：{e}")
        
        return None
    
    def _parse_excel_data(self, excel_file: Path) -> Dict[str, Any]:
        """解析 Excel 文件为数据结构"""
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        
        data = {
            "file": str(excel_file),
            "fields": [],
            "rows": []
        }
        
        # 解析字段定义
        var_row = None
        type_row = None
        desc_rows = []
        data_start_row = None
        
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row[0] == "##var":
                var_row = row
            elif row[0] == "##type":
                type_row = row
            elif row[0] == "##":
                desc_rows.append(row)
            elif var_row and type_row:
                # 第一个非特殊行是数据开始
                data_start_row = i
                break
        
        if var_row and type_row:
            # 提取字段信息
            for j in range(1, len(var_row)):
                field_name = var_row[j] if j < len(var_row) else None
                field_type = type_row[j] if j < len(type_row) else None
                
                if field_name:
                    # 合并所有描述行的注释
                    comments = []
                    for desc_row in desc_rows:
                        if j < len(desc_row) and desc_row[j]:
                            comments.append(str(desc_row[j]))
                    
                    data["fields"].append({
                        "name": field_name,
                        "type": field_type,
                        "comment": " ".join(comments) if comments else ""
                    })
        
        # 解析数据行
        if data_start_row and data["fields"]:
            for i, row in enumerate(sheet.iter_rows(min_row=data_start_row, values_only=True), data_start_row):
                # 跳过空行
                if all(c is None for c in row):
                    continue
                
                row_data = {}
                for j, field in enumerate(data["fields"], 1):
                    value = row[j] if j < len(row) else None
                    if value is not None:
                        row_data[field["name"]] = value
                
                if row_data:
                    data["rows"].append(row_data)
        
        wb.close()
        return data
    
    # ==================== 缓存操作 ====================
    
    def get_file_hash(self, file_path: Path) -> str:
        """计算文件哈希"""
        if not file_path.exists():
            return ""
        
        with open(file_path, "rb") as f:
            return hashlib.md5(f.read()).hexdigest()
    
    def build_cache(self) -> bool:
        """构建缓存"""
        cache_file = self.cache_dir / "config_cache.json"
        
        cache_data = {
            "enums": self.list_enums(),
            "beans": self.list_beans(),
            "tables": self.list_tables(),
            "hashes": {
                "enums": self.get_file_hash(self.enums_file),
                "beans": self.get_file_hash(self.beans_file),
                "tables": self.get_file_hash(self.tables_file)
            }
        }
        
        with open(cache_file, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
        
        print(f"✓ 已构建缓存: {cache_file}")
        return True
    
    def clear_cache(self) -> bool:
        """清除缓存"""
        import shutil
        if self.cache_dir.exists():
            shutil.rmtree(self.cache_dir)
            self.cache_dir.mkdir(exist_ok=True)
        print("✓ 已清除缓存")
        return True


def main():
    parser = argparse.ArgumentParser(description="Luban 配置编辑器辅助脚本")
    parser.add_argument("--data-dir", default="DataTables/Datas", help="数据目录路径")
    
    subparsers = parser.add_subparsers(dest="command", help="命令")
    
    # 枚举命令
    enum_parser = subparsers.add_parser("enum", help="枚举操作")
    enum_subparsers = enum_parser.add_subparsers(dest="enum_command")
    
    enum_list = enum_subparsers.add_parser("list", help="列出所有枚举")
    enum_get = enum_subparsers.add_parser("get", help="获取枚举详情")
    enum_get.add_argument("name", help="枚举名称")
    enum_add = enum_subparsers.add_parser("add", help="新增枚举")
    enum_add.add_argument("name", help="枚举全名 (如 test.EWeaponType)")
    enum_add.add_argument("--values", required=True, help="枚举值，格式: name1=value1:alias1,name2=value2:alias2")
    enum_add.add_argument("--comment", default="", help="枚举注释")
    enum_add.add_argument("--flags", action="store_true", help="是否为标志枚举")
    enum_delete = enum_subparsers.add_parser("delete", help="删除枚举")
    enum_delete.add_argument("name", help="枚举名称")
    
    # Bean 命令
    bean_parser = subparsers.add_parser("bean", help="Bean 操作")
    bean_subparsers = bean_parser.add_subparsers(dest="bean_command")
    
    bean_list = bean_subparsers.add_parser("list", help="列出所有 Bean")
    bean_get = bean_subparsers.add_parser("get", help="获取 Bean 详情")
    bean_get.add_argument("name", help="Bean 名称")
    bean_add = bean_subparsers.add_parser("add", help="新增 Bean")
    bean_add.add_argument("name", help="Bean 全名")
    bean_add.add_argument("--fields", required=True, help="字段定义，格式: name1:type1:comment1,name2:type2:comment2")
    bean_add.add_argument("--parent", default="", help="父类名称")
    bean_add.add_argument("--comment", default="", help="Bean 注释")
    bean_delete = bean_subparsers.add_parser("delete", help="删除 Bean")
    bean_delete.add_argument("name", help="Bean 名称")
    
    # 表命令
    table_parser = subparsers.add_parser("table", help="表操作")
    table_subparsers = table_parser.add_subparsers(dest="table_command")
    table_list = table_subparsers.add_parser("list", help="列出所有表")
    table_get = table_subparsers.add_parser("get", help="获取表数据")
    table_get.add_argument("name", help="表名称")
    
    # 缓存命令
    cache_parser = subparsers.add_parser("cache", help="缓存操作")
    cache_subparsers = cache_parser.add_subparsers(dest="cache_command")
    cache_build = cache_subparsers.add_parser("build", help="构建缓存")
    cache_clear = cache_subparsers.add_parser("clear", help="清除缓存")
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        return
    
    helper = LubanConfigHelper(args.data_dir)
    
    # 枚举操作
    if args.command == "enum":
        if args.enum_command == "list":
            enums = helper.list_enums()
            print(json.dumps(enums, ensure_ascii=False, indent=2))
        elif args.enum_command == "get":
            enum = helper.get_enum(args.name)
            if enum:
                print(json.dumps(enum, ensure_ascii=False, indent=2))
            else:
                print(f"未找到枚举: {args.name}")
        elif args.enum_command == "add":
            # 解析枚举值
            items = []
            for item_str in args.values.split(","):
                parts = item_str.split("=")
                name = parts[0]
                value_alias = parts[1].split(":") if len(parts) > 1 else ["", ""]
                value = value_alias[0] if value_alias[0].isdigit() else None
                alias = value_alias[1] if len(value_alias) > 1 else ""
                if value is None:
                    value = len(items) + 1
                items.append({"name": name, "value": int(value), "alias": alias})
            helper.add_enum(args.name, items, args.flags, True, args.comment)
        elif args.enum_command == "delete":
            helper.delete_enum(args.name)
    
    # Bean 操作
    elif args.command == "bean":
        if args.bean_command == "list":
            beans = helper.list_beans()
            print(json.dumps(beans, ensure_ascii=False, indent=2))
        elif args.bean_command == "get":
            bean = helper.get_bean(args.name)
            if bean:
                print(json.dumps(bean, ensure_ascii=False, indent=2))
            else:
                print(f"未找到 Bean: {args.name}")
        elif args.bean_command == "add":
            # 解析字段
            fields = []
            for field_str in args.fields.split(","):
                parts = field_str.split(":")
                fields.append({
                    "name": parts[0],
                    "type": parts[1] if len(parts) > 1 else "",
                    "comment": parts[2] if len(parts) > 2 else ""
                })
            helper.add_bean(args.name, fields, args.parent, args.comment)
        elif args.bean_command == "delete":
            helper.delete_bean(args.name)
    
    # 表操作
    elif args.command == "table":
        if args.table_command == "list":
            tables = helper.list_tables()
            print(json.dumps(tables, ensure_ascii=False, indent=2))
        elif args.table_command == "get":
            table_data = helper.get_table_data(args.name)
            if table_data:
                print(json.dumps(table_data, ensure_ascii=False, indent=2))
            else:
                print(f"未找到表：{args.name}")
    
    # 缓存操作
    elif args.command == "cache":
        if args.cache_command == "build":
            helper.build_cache()
        elif args.cache_command == "clear":
            helper.clear_cache()


if __name__ == "__main__":
    main()
